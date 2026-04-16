"""
Report Renderers — Strategy pattern for report output formats.

BaseRenderer defines the contract; PdfRenderer and ExcelRenderer are
concrete strategies. Add new formats (CSV, HTML) by subclassing.
"""
import io
import logging
from abc import ABC, abstractmethod
from datetime import datetime

import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl.styles import Font, PatternFill, Alignment


log = logging.getLogger(__name__)

HEADER_COLOR = colors.HexColor("#1a365d")
ALT_ROW_COLOR = colors.HexColor("#f7fafc")
WHITE = colors.white


def _table_style(col_count: int) -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [ALT_ROW_COLOR, WHITE]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])


class BaseRenderer(ABC):
    """Abstract report renderer (Strategy interface)."""

    @abstractmethod
    def render(self, sections: list[dict], title: str, **kwargs) -> io.BytesIO:
        """Render named sections into an output buffer.

        Each section dict has keys:
          - heading: str
          - dataframe: pd.DataFrame  (may be empty)
          - columns: list[str]       (display columns)
        """
        ...

    @abstractmethod
    def content_type(self) -> str:
        ...

    @abstractmethod
    def file_extension(self) -> str:
        ...


class PdfRenderer(BaseRenderer):
    """Renders report sections as a multi-section PDF."""

    def render(self, sections: list[dict], title: str, **kwargs) -> io.BytesIO:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                leftMargin=40, rightMargin=40,
                                topMargin=50, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements: list = []

        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]
        ))
        elements.append(Spacer(1, 20))

        for section in sections:
            elements.append(Paragraph(section["heading"], styles["Heading2"]))
            df: pd.DataFrame = section["dataframe"]
            cols = section.get("columns") or list(df.columns)

            if df.empty:
                elements.append(Paragraph("No data found.", styles["Normal"]))
            else:
                data = [cols] + df[cols].values.tolist()
                t = Table(data, colWidths=[max(80, 400 // len(cols))] * len(cols))
                t.setStyle(_table_style(len(cols)))
                elements.append(t)
            elements.append(Spacer(1, 20))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def content_type(self) -> str:
        return "application/pdf"

    def file_extension(self) -> str:
        return "pdf"


class ExcelRenderer(BaseRenderer):
    """Renders report sections as multi-sheet Excel workbook."""

    def render(self, sections: list[dict], title: str, **kwargs) -> io.BytesIO:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for section in sections:
                df: pd.DataFrame = section["dataframe"]
                sheet = section.get("sheet_name", section["heading"][:31])

                if df.empty:
                    pd.DataFrame([{"message": "No data found"}]).to_excel(
                        writer, sheet_name=sheet, index=False
                    )
                else:
                    df.to_excel(writer, sheet_name=sheet, index=False)
                    ws = writer.sheets[sheet]
                    header_fill = PatternFill(
                        start_color="1a365d", end_color="1a365d", fill_type="solid"
                    )
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                        ws.column_dimensions[cell.column_letter].width = 20

        output.seek(0)
        return output

    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def file_extension(self) -> str:
        return "xlsx"

